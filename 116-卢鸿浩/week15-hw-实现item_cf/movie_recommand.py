import openpyxl
import numpy as np
import time
from collections import defaultdict

'''
电影打分数据集
实现协同过滤

excel读取
cos距离计算
排序

'''

#为了好理解，将数据格式转化成user-item的打分矩阵形式
def build_u2i_matrix(user_item_score_data_path, item_name_data_path, write_file=False):
    #获取item id到电影名的对应关系
    item_id_to_item_name = {}
    with open(item_name_data_path, encoding="ISO-8859-1") as f:
        for line in f:
            item_id, item_name = line.split("|")[:2]
            item_id = int(item_id)
            item_id_to_item_name[item_id] = item_name
    total_movie_count = len(item_id_to_item_name)
    print("total movie:", total_movie_count)

    #读打分文件
    user_to_rating = {}
    with open(user_item_score_data_path, encoding="ISO-8859-1") as f:
        for line in f:
            user_id, item_id, score, time_stamp = line.split("\t")
            user_id, item_id, score = int(user_id), int(item_id), int(score)
            if user_id not in user_to_rating:
                user_to_rating[user_id] = [0] * total_movie_count
            user_to_rating[user_id][item_id - 1] = score
    print("total user:", len(user_to_rating))

    if not write_file:
        return user_to_rating, item_id_to_item_name

    # 写入excel便于查看
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(index=0)
    sheet_1 = workbook.create_sheet(index=1)
    #第一行：user_id, movie1, movie2...
    header = ["user_id"] + [item_id_to_item_name[i + 1] for i in range(total_movie_count)]
    sheet.append(header)
    for i in range(len(user_to_rating)):
        #每行：user_id, rate1, rate2...
        line = [i + 1] + user_to_rating[i + 1]
        sheet.append(line)
    workbook.save("user_movie_rating.xlsx")
    return user_to_rating, item_id_to_item_name

#向量余弦距离
def cosine_distance(vector1, vector2):
    ab = vector1.dot(vector2)
    a_norm = np.sqrt(np.sum(np.square(vector1)))
    b_norm = np.sqrt(np.sum(np.square(vector2)))
    return ab/(a_norm * b_norm)

#根据用户打分计算item相似度
def find_similar_item(user_to_rating):
    item_to_vector = {}  # {item_id, [score1, score2,...]}
    total_user = len(user_to_rating)
    for user, user_rating in user_to_rating.items():
        #for moive_id, score in enumerate(user_to_rating):  # 原始版本
        for moive_id, score in enumerate(user_rating):
            #print(moive_id, score)
            moive_id += 1
            if moive_id not in item_to_vector:
                item_to_vector[moive_id] = [0] * (total_user + 1)
            item_to_vector[moive_id][user] = score
    #item_to_vector记录了每个用户打分，数据结构和user_to_rating一样
    #复用一下下方的相似度计算方法
    return find_similar_user(item_to_vector)

#依照user对item的打分判断user之间的相似度
def find_similar_user(user_to_rating):  # user_to_rating: {user_id, [score1, score2, ...]}
    user_to_similar_user = {}   # {user_ida, [[user_idb,similarity],[user_idc, similarity], ...]}
    score_buffer = {}           # {"user_ida_user_idb", similarity}
    for user_a, ratings_a in user_to_rating.items():
        similar_user = []
        for user_b, ratings_b in user_to_rating.items():
            # #全算比较慢，省去一部分用户
            if user_b == user_a:  #  or user_b > 100 or user_a > 100:
                continue
            #ab用户互换不用重新计算cos
            if "%d_%d"%(user_b, user_a) in score_buffer:
                similarity = score_buffer["%d_%d"%(user_b, user_a)]
            #相似度计算采取cos距离
            else:
                similarity = cosine_distance(np.array(ratings_a), np.array(ratings_b))
                score_buffer["%d_%d" % (user_a, user_b)] = similarity
            similar_user.append([user_b, similarity])
        similar_user = sorted(similar_user, reverse=True, key=lambda x:x[1])
        user_to_similar_user[user_a] = similar_user
    return user_to_similar_user

#基于user的协同过滤
#输入user_id, item_id, 给出预测打分
#有预测打分之后就可以对该用户所有未看过的电影打分，然后给出排序结果
#所以实现打分函数即可
#topn为考虑多少相似的用户
#取前topn相似用户对该电影的打分
def user_cf(user_id, item_id, user_to_similar_user, user_to_rating, topn=10):
    pred_score = 0
    count = 0
    for similar_user, similarity in user_to_similar_user[user_id][:topn]:
        #相似用户对这部电影的打分
        rating_by_similiar_user = user_to_rating[similar_user][item_id - 1]
        #分数*用户相似度，作为一种对分数的加权，越相似的用户评分越重要
        pred_score += rating_by_similiar_user * similarity
        #如果这个相似用户没看过该电影，就不计算在总数内
        if rating_by_similiar_user != 0:
            count += 1
    pred_score /= count + 1e-5
    return pred_score


#基于item的协同过滤
#类似user_cf
#自己尝试实现
def item_cf(user_id, item_id, similar_items, user_to_rating, topn = 5):
    pred_score = 0
    count = 0
    count_similar = 0
    pred_score_similar = 0
    print("/////////////////////////////")
    print(similar_items[item_id - 1])
    for similar_items, similarity in similar_items[item_id - 1][:topn]:
        print("\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\")
        print(item_id, similar_items, similarity)
        # 相似电影的得分
        rating_by_similar_items = user_to_rating[user_id][similar_items - 1]
        # 分数*电影相似度，作为分数的加权，越相似的电影评分越重要
        pred_score += rating_by_similar_items*similarity
        # 如果这个相似电影没有被该用户评分，就不计算在总数内
        if rating_by_similar_items != 0:
            count += 1
            count_similar += similarity
    pred_score_similar = pred_score / (count_similar + 1e-5)
    pred_score /= count + 1e-5
    return pred_score, pred_score_similar


#对于一个用户做完整的item召回
def movie_recommand(user_id, similar_user, similar_items, user_to_rating, item_to_name, topn=10):
    #当前用户还没看过的所有电影id
    unseen_items = [item_id + 1 for item_id, rating in enumerate(user_to_rating[user_id]) if rating == 0]
    res = []
    res2 = []
    for item_id in unseen_items:
        #user_cf打分
        #score = user_cf(user_id, item_id, similar_user, user_to_rating)
        # item_cf打分
        score, score2 = item_cf(user_id, item_id, similar_items, user_to_rating)
        res.append([item_to_name[item_id], score])
        res2.append([item_to_name[item_id], score2])
    #排序输出
    res = sorted(res, key=lambda x:x[1], reverse=True)
    res2 = sorted(res2, key=lambda x: x[1], reverse=True)
    return res[:topn], res2[:topn]



if __name__ == "__main__":
    user_item_score_data_path = "ml-100k/u.data"
    item_name_data_path = "ml-100k/u.item"
    user_to_rating, item_to_name = build_u2i_matrix(user_item_score_data_path, item_name_data_path, False)

    #user-cf
    similar_user = find_similar_user(user_to_rating)
    similar_items = find_similar_item(user_to_rating)
    print("--------")
    print(similar_items[3])
    print(item_to_name[3])
    print(item_to_name[similar_items[3][0][0]])
    print("========")
    print(similar_user[1])
    # print("相似用户计算完成，耗时：", time.time() - s)
    # while True:
    #     user_id = int(input("输入用户id："))
    #     item_id = int(input("输入电影id："))
    #     res = user_cf(user_id, item_id, similar_user, user_to_rating)
    #     print(res)

    #为用户推荐电影
    while True:
        user_id = int(input("输入用户id："))
        recommands, recommands2 = movie_recommand(user_id, similar_user, similar_items, user_to_rating, item_to_name)
        for recommand, score in recommands:
            print("%.4f\t%s"%(score, recommand))
        print("2222222222222222")
        for recommand, score in recommands2:
            print("%.4f\t%s"%(score, recommand))
